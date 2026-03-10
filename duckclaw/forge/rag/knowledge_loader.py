"""
KnowledgeLoader — pipeline de ingesta de catálogo a DuckDB con embeddings.

Spec: specs/DuckDB_Native_RAG_Vector_Search.md

Uso:
  python -m duckclaw.forge.rag.knowledge_loader /path/to/catalogo.csv
  python -m duckclaw.forge.rag.knowledge_loader /path/to/catalogo.xlsx --schema powerseal_worker
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from pathlib import Path


def _get_db_path() -> str:
    p = os.environ.get("DUCKCLAW_DB_PATH", "").strip()
    if p:
        return str(Path(p).resolve())
    return str(Path(__file__).resolve().parents[3] / "db" / "workers.duckdb")


def load_csv(path: str) -> list[dict]:
    """Carga CSV con columnas: sku, name, description, price, stock_status."""
    rows = []
    with open(path, encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append({k.strip(): v.strip() if isinstance(v, str) else v for k, v in r.items()})
    return rows


def load_xlsx(path: str) -> list[dict]:
    """Carga primera hoja de XLSX."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2):
            r = dict(zip(headers, [str(c.value or "").strip() for c in row]))
            if any(r.values()):
                rows.append(r)
        return rows
    except ImportError:
        return []


def run(path: str, schema: str = "powerseal_worker", db_path: str | None = None) -> int:
    """Ingesta catálogo desde CSV o XLSX. Retorna número de filas insertadas."""
    p = Path(path)
    if not p.is_file():
        print(f"Archivo no encontrado: {path}", file=sys.stderr)
        return 0

    if p.suffix.lower() == ".csv":
        rows = load_csv(str(p))
    elif p.suffix.lower() in (".xlsx", ".xls"):
        rows = load_xlsx(str(p))
    else:
        rows = load_csv(str(p))

    if not rows:
        print("No hay filas para procesar.", file=sys.stderr)
        return 0

    # Normalizar columnas (case-insensitive)
    def norm(r):
        return {k.lower().replace(" ", "_"): v for k, v in r.items()}

    rows = [norm(r) for r in rows]
    keys = set()
    for r in rows:
        keys.update(r.keys())

    sku_col = next((k for k in keys if k in ("sku", "id", "codigo")), list(keys)[0])
    name_col = next((k for k in keys if k in ("name", "nombre", "producto")), "name")
    desc_col = next((k for k in keys if k in ("description", "descripcion")), "description")
    price_col = next((k for k in keys if k in ("price", "precio")), "price")
    stock_col = next((k for k in keys if k in ("stock_status", "stock", "disponibilidad")), "stock_status")

    from duckclaw import DuckClaw
    from duckclaw.forge.rag import ensure_catalog_schema, embed_text

    db = DuckClaw(db_path or _get_db_path())
    ensure_catalog_schema(db, schema)

    embedding_fn = embed_text
    inserted = 0

    for r in rows:
        sku = str(r.get(sku_col, "") or "").strip()[:64]
        name = str(r.get(name_col, "") or "").strip()[:512]
        desc = str(r.get(desc_col, "") or "").strip()[:2048]
        price_val = r.get(price_col, "") or "0"
        try:
            price = float(str(price_val).replace(",", "."))
        except ValueError:
            price = 0.0
        stock = str(r.get(stock_col, "") or "").strip()[:64]

        if not sku:
            continue

        text_to_embed = f"{name} {desc}".strip()
        emb = embedding_fn(text_to_embed) if text_to_embed else None

        if emb:
            vec_str = "[" + ",".join(str(x) for x in emb) + "]"
            sku_esc = sku.replace("'", "''")
            name_esc = name.replace("'", "''")
            desc_esc = desc.replace("'", "''")
            stock_esc = stock.replace("'", "''")
            try:
                db.execute(
                    f"""
                    INSERT INTO {schema}.catalog_items (sku, name, description, price, stock_status, embedding)
                    VALUES ('{sku_esc}', '{name_esc}', '{desc_esc}', {price}, '{stock_esc}', {vec_str}::FLOAT[384])
                    ON CONFLICT (sku) DO UPDATE SET
                        name = EXCLUDED.name,
                        description = EXCLUDED.description,
                        price = EXCLUDED.price,
                        stock_status = EXCLUDED.stock_status,
                        embedding = EXCLUDED.embedding
                    """
                )
                inserted += 1
            except Exception as e:
                print(f"Error insertando {sku}: {e}", file=sys.stderr)
        else:
            # Sin embeddings: insertar sin vector (catalog_retriever usará fallback LIKE)
            sku_esc = sku.replace("'", "''")
            name_esc = name.replace("'", "''")
            desc_esc = desc.replace("'", "''")
            stock_esc = stock.replace("'", "''")
            try:
                db.execute(
                    f"""
                    INSERT INTO {schema}.catalog_items (sku, name, description, price, stock_status)
                    VALUES ('{sku_esc}', '{name_esc}', '{desc_esc}', {price}, '{stock_esc}')
                    ON CONFLICT (sku) DO UPDATE SET
                        name = EXCLUDED.name,
                        description = EXCLUDED.description,
                        price = EXCLUDED.price,
                        stock_status = EXCLUDED.stock_status
                    """
                )
                inserted += 1
            except Exception as e:
                print(f"Error insertando {sku}: {e}", file=sys.stderr)

    return inserted


def main() -> None:
    parser = argparse.ArgumentParser(description="KnowledgeLoader: ingesta catálogo a DuckDB RAG")
    parser.add_argument("path", help="Ruta a catalogo.csv o catalogo.xlsx")
    parser.add_argument("--schema", default="powerseal_worker", help="Schema DuckDB")
    parser.add_argument("--db-path", default=None, help="Ruta a .duckdb")
    args = parser.parse_args()
    n = run(args.path, schema=args.schema, db_path=args.db_path)
    print(f"Ingestadas {n} filas.")


if __name__ == "__main__":
    main()
