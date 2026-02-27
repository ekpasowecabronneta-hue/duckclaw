"""
Exportación de consultas SQL a Excel (.xlsx) descargables.

Uso:
    from duckclaw.bi.excel_export import export_query_to_excel
    path = export_query_to_excel(db, "SELECT * FROM olist_orders LIMIT 100", save_dir="output")
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any


def export_query_to_excel(
    db: Any,
    sql: str,
    save_dir: str = "output",
    sheet_name: str = "Datos",
    limit: int = 10000,
) -> str:
    """
    Ejecuta una consulta SQL y guarda el resultado en un archivo Excel (.xlsx).

    - sql: consulta SELECT (solo SELECT/WITH permitido). Incluir LIMIT si quieres menos filas.
    - save_dir: directorio donde guardar el archivo.
    - sheet_name: nombre de la hoja en Excel (máx 31 caracteres).
    - limit: máximo de filas a exportar (por defecto 10000).

    Devuelve la ruta absoluta del archivo generado o un mensaje de error.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "Error: instala openpyxl (pip install openpyxl) para exportar a Excel."

    sql_upper = (sql or "").strip().upper()
    if not (sql_upper.startswith("SELECT") or sql_upper.startswith("WITH")):
        return "Error: solo se permiten consultas SELECT o WITH."
    for blocked in ("DROP", "INSERT", "UPDATE", "DELETE", "ALTER", "CREATE", "TRUNCATE"):
        if blocked in sql_upper:
            return f"Error: no se permite {blocked} en la consulta."

    try:
        raw = db.query(sql)
        rows = json.loads(raw) if isinstance(raw, str) else (raw or [])
    except Exception as e:
        return f"Error ejecutando SQL: {e}"

    if not rows:
        return "No hay datos para exportar."

    # Limitar filas
    rows = rows[: int(limit)]

    # Asegurar directorio
    out_dir = Path(save_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    # Nombre de archivo único por hash del SQL
    sig = hashlib.md5(sql.encode()).hexdigest()[:8]
    filename = f"export_{sig}.xlsx"
    out_path = out_dir / filename

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet_name_safe = (sheet_name or "Datos")[:31].replace(":", "").replace("\\", "").replace("/", "").replace("*", "").replace("?", "").replace("[", "").replace("]", "")
    ws.title = sheet_name_safe or "Datos"

    # Encabezados
    cols = list(rows[0].keys()) if rows else []
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = openpyxl.styles.Font(bold=True)

    # Filas
    for r, row in enumerate(rows, 2):
        for c, col in enumerate(cols, 1):
            val = row.get(col)
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # Ajustar anchos de columna
    for c in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(20, max(10, len(str(cols[c - 1])) + 2))

    wb.save(out_path)
    return f"Archivo Excel guardado: {out_path}"
